from flask import Blueprint, flash, request, redirect, url_for, render_template, Response, current_app, jsonify, session
from flask_login import login_required, current_user
import multiprocessing
from datetime import datetime
from sqlalchemy.sql import func

from . import db, login_manager, socketio
from .models_db import RoadDamage, RoadData, Users
import uuid
from apps.libserver.lib import compress_video, get_province_name, restart_server
from apps.libserver.query_lib import single_road_data, road_datas, get_sort_road_data, get_all_damage_province


from openpyxl import Workbook
from openpyxl.styles import Alignment
from io import BytesIO
from openpyxl.utils import get_column_letter

from threading import Thread

import json

from apps.libyolov.mainstream import StreamMainDetect



detect = Blueprint('detect', __name__)


main_detect_stream = StreamMainDetect()

@detect.route('/check_session')
def check_session():
    # Memeriksa apakah ada data 'username' dalam sesi
    if session.get('road_data') is not None:
        print(session.get('road_data'))
        return 'Sesi sudah masuk'
    else:
        return 'Sesi belum masuk'

@detect.route('/streamcam', methods=['GET','POST'])
@login_required
def streamcam():
    if request.method == 'POST':
        road_data = RoadData()
        road_data.id = str(uuid.uuid4())
        road_data.province = request.form.get('province')
        road_data.city = request.form.get('city')
        road_data.street_name = request.form.get('streetname')
        road_data.user_id = current_user.id


        port_camera = request.form.get('port_camera')
      
        
        road_data_dict = {
            'id': road_data.id,
            'province': road_data.province,
            'city': road_data.city,
            'street_name': road_data.street_name,
            'user_id': road_data.user_id
        }

        road_data_json = json.dumps(road_data_dict)

        session['road_data'] = road_data_json

        db.session.add(road_data)
        db.session.commit()

        main_detect_stream.start_run(road_data.id, port_camera, socketio)

        flash('The process is loading, please wait..', 'success')

        return render_template('home/output.html', 
                        segment='streamcam',
                        data=road_data,
                        get_province_name=get_province_name,
                        user_id=current_user.id)
    
    road_data_session = session.get('road_data')
    
    if road_data_session:
        road_data_dict = json.loads(road_data_session)
        road_data = RoadData(**road_data_dict)

        return render_template('home/output.html', 
                            segment='streamcam',
                            data=road_data,
                            get_province_name=get_province_name,
                            user_id=current_user.id)
    return redirect(url_for('views.index'))

@detect.route('/stopcam', methods=['POST'])
def stopcam():
    if request.method=='POST':
        main_detect_stream.stop_run()
        flash('Congratulations, your detection finished!', 'success')
        session.pop('road_data', None)
        return redirect(url_for('views.report'))


@detect.route('/result/<id>')
@login_required
def result(id):
    check_id = RoadData.query.get(id)
    if check_id is None:
        # Jika ID tidak ada dalam database, arahkan ke halaman beranda
        flash('Id not found.', 'danger')
        return redirect(url_for('views.index'))
    
    road_data = single_road_data(id, current_user.id)

    data = [ 
            ['Type Damage', 'Total'], 
            ['Pathole', int(road_data.pathole)], 
            ['Crocodile', int(road_data.crocodile)], 
            ['Longitudinal', int(road_data.longitudinal)], 
            ['Transversal', int(road_data.transversal)] 
        ]
    return render_template('home/result.html', 
                            segment='report',
                            road_data=road_data,
                            get_province_name=get_province_name,
                            data=data,
                            user_id=current_user.id)

@detect.route('/download_excel')
@login_required
def download_excel():
    wb = Workbook()
    ws = wb.active

    header = [
        'No', 'Time/Date', 'Province', 'City', 'Street Name',
        'Pathole', 'Alligator', 'Longitudinal', 'Transversal',
    ]

    data = road_datas(current_user.id)
    ws.append(header)

    # Menambahkan data ke lembar kerja
    for i, row in enumerate(data, start=1):
        row = [i, row[1], get_province_name(row[2]), row[3], row[4], row[5], row[6], row[7], row[8],]  # Menambahkan kolom 'Action' dengan string kosong
        ws.append(row)

    # Mengukur lebar kolom untuk header
    for column_num, column_title in enumerate(header, 1):
        column_letter = get_column_letter(column_num)
        max_length = max(len(str(column_title)), max(len(str(value)) for value in ws[column_letter][1:]) + 2)
        ws.column_dimensions[column_letter].width = max_length

    # Mengukur lebar kolom untuk data
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        for column_num, value in enumerate(row, 1):
            column_letter = get_column_letter(column_num)
            max_length = max(max(len(str(value)) for value in ws[column_letter]), 2)
            ws.column_dimensions[column_letter].width = max_length

    # Mengatur alignment (penyelarasan) untuk header
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    current_time = datetime.now().strftime("%Y%m%d%H%M%S")

    filename = f"road_data_{current_time}.xlsx"

    return Response(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment;filename={filename}"})


@detect.route('/update_graph')
@login_required
def update_graph():
    selected_province = request.args.get('province')

    if not selected_province:

        selected_province = 0

    data_label = get_all_damage_province(selected_province)

    return jsonify(data_label)


@detect.route('/update_chart')
@login_required
def update_chart():
    sort_option = request.args.get('sort_option')

    data = get_sort_road_data(sort_option, current_user.id)

    return jsonify(data)


